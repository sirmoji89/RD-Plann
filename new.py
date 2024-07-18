import openpyxl
import os
from datetime import datetime
import xml.etree.ElementTree as ET

# Initialize global variables
project_data = {}
personnel_data = []

def read_project_data(file_name):
    global project_data, personnel_data

    # Load the Excel file
    wb = openpyxl.load_workbook(file_name)

    # Select the "Projects List" sheet
    projects_sheet = wb['Projects List']

    # Initialize dictionary to store project data
    project_data = {}

    # Iterate over rows starting from row 6
    for row in range(6, projects_sheet.max_row + 1):
        # Get the project name and code
        project_name = projects_sheet.cell(row=row, column=1).value
        project_code = str(projects_sheet.cell(row=row, column=2).value).zfill(4)

        # Check if the project name is not blank
        if project_name:
            # Append to the dictionary
            project_data[project_name] = {
                'Project Code': project_code,
                'Phases': {}
            }

    # Select the "Personnel List" sheet
    personnel_sheet = wb['Personnel List']

    # Iterate over rows starting from row 6
    for row in range(6, personnel_sheet.max_row + 1):
        # Get the personnel name
        personnel_name = personnel_sheet.cell(row=row, column=1).value

        # Check if the personnel name is not blank
        if personnel_name:
            # Append to the list
            personnel_data.append(personnel_name)

    # Close the workbook
    wb.close()

    # Outputs
    return project_data, personnel_data


def read_wbs_files(project_data):
    for project, data in project_data.items():
        project_code = data['Project Code']
        file_name = f"RD-{project_code}-WBS.xlsx"

        # Check if the file exists
        if os.path.isfile(file_name):
            # Open the workbook
            wb = openpyxl.load_workbook(file_name)

            # Select the "WBS" sheet
            wbs_sheet = wb['WBS']

            # Iterate over rows starting from row 7
            current_phase = None
            for row in range(7, wbs_sheet.max_row + 1):
                # Get the value in column A
                phase_name = wbs_sheet.cell(row=row, column=1).value
                if phase_name:
                    # Phase row
                    phase_number = wbs_sheet.cell(row=row, column=2).value
                    responsible_person = wbs_sheet.cell(row=row, column=3).value
                    date_of_finishing = wbs_sheet.cell(row=row, column=8).value
                    required_time = wbs_sheet.cell(row=row, column=9).value

                    # Create phase entry
                    current_phase = {
                        'Phase Name': phase_name,
                        'Phase Number': phase_number,
                        'Responsible Person': responsible_person,
                        'Date of Finishing': date_of_finishing,
                        'Required Time': required_time,
                        'Tasks': []
                    }

                    # Add to project data
                    data['Phases'][f"{phase_name} - {phase_number}"] = current_phase
                else:
                    # Task row
                    task_name = wbs_sheet.cell(row=row, column=2).value
                    first_coworker = wbs_sheet.cell(row=row, column=4).value
                    second_coworker = wbs_sheet.cell(row=row, column=5).value
                    third_coworker = wbs_sheet.cell(row=row, column=6).value
                    task_status = wbs_sheet.cell(row=row, column=7).value
                    date_of_finishing_task = wbs_sheet.cell(row=row, column=8).value
                    required_time_task = wbs_sheet.cell(row=row, column=9).value

                    task_data = {
                        'Task Name': task_name,
                        'First Coworker': first_coworker,
                        'Second Coworker': second_coworker,
                        'Third Coworker': third_coworker,
                        'Task Status': task_status,
                        'Date of Finishing': date_of_finishing_task,
                        'Required Time': required_time_task
                    }

                    # Add to current phase
                    if current_phase:
                        current_phase['Tasks'].append(task_data)

            # Close the workbook
            wb.close()
        else:
            print(f"File not found: {file_name}")

    return project_data


def update_active_phases(file_name, project_data):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_name)

    # Select the "Active Phases" sheet
    active_phases_sheet = wb['Active Phases']

    # Iterate over rows starting from row 6
    for row in range(6, active_phases_sheet.max_row + 1):
        # Get the project name, active phase, and active phase number
        project_name = active_phases_sheet.cell(row=row, column=1).value
        active_phase = active_phases_sheet.cell(row=row, column=2).value
        active_phase_number = active_phases_sheet.cell(row=row, column=3).value

        # Check if the project name is not blank
        if project_name and project_name in project_data:
            phase_key = f"{active_phase} - {active_phase_number}"
            if phase_key in project_data[project_name]['Phases']:
                project_data[project_name]['Phases'][phase_key]['IsActive'] = True

    # Close the workbook
    wb.close()


def read_timesheet(file_name, timesheet_data, personnel_name):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_name)

    # Select the "Sheet1" sheet
    sheet = wb['Sheet1']

    # Get the dates from row 4
    dates = []
    for col in range(5, sheet.max_column + 1):
        date = sheet.cell(row=4, column=col).value
        if isinstance(date, datetime):
            dates.append(date)

    # Iterate over rows starting from row 8
    for row in range(8, sheet.max_row + 1):
        project_name = sheet.cell(row=row, column=1).value
        phase_name = sheet.cell(row=row, column=2).value
        phase_number = sheet.cell(row=row, column=3).value
        task_name = sheet.cell(row=row, column=4).value

        if project_name and phase_name and phase_number and task_name:
            task_key = (project_name, phase_name, phase_number, task_name)

            if task_key not in timesheet_data:
                timesheet_data[task_key] = {}

            if personnel_name not in timesheet_data[task_key]:
                timesheet_data[task_key][personnel_name] = []

            for col, date in enumerate(dates, start=5):
                hours = sheet.cell(row=row, column=col).value
                if hours:
                    timesheet_data[task_key][personnel_name].append((date, hours))

    # Close the workbook
    wb.close()


def write_to_xml(file_name, project_data, timesheet_data):
    root = ET.Element("Projects")

    for project_name, project_info in project_data.items():
        project_elem = ET.SubElement(root, "Project", name=project_name, code=project_info['Project Code'])

        for phase_key, phase_info in project_info['Phases'].items():
            phase_elem = ET.SubElement(project_elem, "Phase", name=phase_info['Phase Name'], number=str(phase_info['Phase Number']))
            ET.SubElement(phase_elem, "ResponsiblePerson").text = str(phase_info['Responsible Person']) if phase_info['Responsible Person'] else ''
            ET.SubElement(phase_elem, "DateOfFinishing").text = str(phase_info['Date of Finishing']) if phase_info['Date of Finishing'] else ''
            ET.SubElement(phase_elem, "RequiredTime").text = str(phase_info['Required Time']) if phase_info['Required Time'] else ''
            ET.SubElement(phase_elem, "IsActive").text = str(phase_info.get('IsActive', False))

            for task_info in phase_info['Tasks']:
                task_elem = ET.SubElement(phase_elem, "Task", name=task_info['Task Name'])
                ET.SubElement(task_elem, "FirstCoworker").text = str(task_info['First Coworker']) if task_info['First Coworker'] else ''
                ET.SubElement(task_elem, "SecondCoworker").text = str(task_info['Second Coworker']) if task_info['Second Coworker'] else ''
                ET.SubElement(task_elem, "ThirdCoworker").text = str(task_info['Third Coworker']) if task_info['Third Coworker'] else ''
                ET.SubElement(task_elem, "TaskStatus").text = str(task_info['Task Status']) if task_info['Task Status'] else ''
                ET.SubElement(task_elem, "DateOfFinishing").text = str(task_info['Date of Finishing']) if task_info['Date of Finishing'] else ''
                ET.SubElement(task_elem, "RequiredTime").text = str(task_info['Required Time']) if task_info['Required Time'] else ''

                task_key = (project_name, phase_info['Phase Name'], phase_info['Phase Number'], task_info['Task Name'])
                if task_key in timesheet_data:
                    for personnel_name, records in timesheet_data[task_key].items():
                        personnel_elem = ET.SubElement(task_elem, "Personnel", name=personnel_name)
                        for date, hours in records:
                            ET.SubElement(personnel_elem, "Record", date=date.strftime("%Y-%m-%d"), hours=str(hours))

    tree = ET.ElementTree(root)
    tree.write(file_name, encoding='utf-8', xml_declaration=True)

# Main program execution
resource_projects_file = 'Resource & Projects.xlsx'
project_data, personnel_data = read_project_data(resource_projects_file)
project_data = read_wbs_files(project_data)
update_active_phases(resource_projects_file, project_data)

# Initialize timesheet data dictionary
timesheet_data = {}

# Read timesheet files for each personnel
for personnel_name in personnel_data:
    timesheet_file = f"Timesheet-{personnel_name}.xlsx"
    if os.path.isfile(timesheet_file):
        read_timesheet(timesheet_file, timesheet_data, personnel_name)

# Write data to XML
output_xml_file = 'output.xml'
write_to_xml(output_xml_file, project_data, timesheet_data)

print(f"Data written to {output_xml_file}")
